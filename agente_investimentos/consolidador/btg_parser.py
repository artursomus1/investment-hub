"""Parser para relatorios do BTG Pactual (XLSX)."""

import re
from datetime import datetime
from typing import List

import openpyxl

from agente_investimentos.consolidador.models import ParsedAssetInst, InstitutionData


def _safe_float(value) -> float:
    """Converte valor para float de forma segura."""
    if value is None or value == "-" or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(".", "").replace(",", ".")
    text = re.sub(r'[^\d.\-]', '', text)
    try:
        return float(text)
    except (ValueError, TypeError):
        return 0.0


def _safe_str(value) -> str:
    """Converte valor para string de forma segura."""
    if value is None:
        return ""
    return str(value).strip()


def _parse_date(value) -> str:
    """Converte data para string."""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if value:
        return str(value)[:10]
    return ""


def _parse_renda_fixa(ws, instituicao: str) -> List[ParsedAssetInst]:
    """Extrai ativos de renda fixa da sheet 'Renda Fixa'.

    Usa apenas a secao 'Posicao' (linhas 6-~10), para antes de 'Posicoes Detalhadas'.
    """
    ativos = []
    in_posicao = False

    for row_idx in range(4, min(50, ws.max_row + 1)):
        b_val = _safe_str(ws.cell(row=row_idx, column=2).value)

        # Detecta inicio da secao de posicoes
        if "Posição >" in b_val or "Posicao >" in b_val:
            in_posicao = True
            continue

        # Para ao encontrar secao de detalhamento ou movimentacao
        if b_val.startswith("Posições Detalhadas") or b_val.startswith("Posicoes Detalhadas"):
            in_posicao = False
            continue
        if b_val.startswith("Moviment"):
            break
        if b_val.startswith("Posição Consolidada"):
            in_posicao = False
            continue

        if not in_posicao:
            continue

        # Ignora headers e totais
        if b_val.lower() in ("emissor", "total", ""):
            continue

        emissor_val = b_val
        ativo_val = _safe_str(ws.cell(row=row_idx, column=3).value)  # C
        saldo_bruto = _safe_float(ws.cell(row=row_idx, column=12).value)  # L

        if saldo_bruto <= 0 or not emissor_val:
            continue

        vencimento = _parse_date(ws.cell(row=row_idx, column=5).value)  # E
        taxa_str = _safe_str(ws.cell(row=row_idx, column=9).value)  # I
        ir = _safe_float(ws.cell(row=row_idx, column=13).value)  # M
        saldo_liq = saldo_bruto - ir

        # Subtipo
        subtipo = "CDB"
        nome_upper = (ativo_val or emissor_val).upper()
        for key in ("CRA", "CRI", "LCA", "LCI", "DEBENTURE", "DEB"):
            if key in nome_upper:
                subtipo = key if key != "DEB" else "Debenture"
                break

        # Indexador
        indexador = ""
        taxa_upper = taxa_str.upper()
        if "CDI" in taxa_upper or "% A.A." in taxa_upper:
            indexador = "CDI"
        elif "IPCA" in taxa_upper:
            indexador = "IPCA"
        elif "PRE" in taxa_upper:
            indexador = "PRE"

        nome = f"{subtipo} {emissor_val}" if ativo_val.startswith("CDB-") else ativo_val or emissor_val

        ativos.append(ParsedAssetInst(
            nome=nome,
            ticker="",
            tipo="Renda Fixa",
            subtipo=subtipo,
            saldo_bruto=saldo_bruto,
            saldo_liquido=saldo_liq,
            impostos=ir,
            alocacao_pct=0.0,
            rent_mes=0.0,
            rent_ano=0.0,
            rent_12m=0.0,
            indexador=indexador,
            taxa=taxa_str,
            vencimento=vencimento,
            emissor=emissor_val,
            instituicao=instituicao,
        ))

    return ativos


def _parse_renda_variavel(ws, instituicao: str) -> List[ParsedAssetInst]:
    """Extrai posicao atual de acoes da sheet 'Renda Variavel'.

    O BTG mostra movimentacoes (compras/vendas), precisamos calcular posicao atual.
    So processa a secao 'Movimentacao > Acoes', para ao chegar em 'Total de Compras'.
    """
    posicoes = {}  # ticker -> {"qtd": int, "custo_total": float}
    in_acoes = False

    for row_idx in range(4, ws.max_row + 1):
        b_val = _safe_str(ws.cell(row=row_idx, column=2).value)

        # Detecta inicio da secao de acoes
        if "Ações" in b_val and "Movimentação" in b_val and "Aluguel" not in b_val:
            in_acoes = True
            continue
        # Headers
        if b_val == "Data":
            continue

        # Para ao chegar nos totais ou secoes de aluguel/opcoes
        if b_val.startswith("Total"):
            if in_acoes:
                break
            continue
        if "Aluguel" in b_val or "Opções" in b_val or "Opcoes" in b_val:
            break

        if not in_acoes:
            continue

        transacao = _safe_str(ws.cell(row=row_idx, column=3).value)  # C
        ticker = _safe_str(ws.cell(row=row_idx, column=4).value)  # D
        qtd = _safe_float(ws.cell(row=row_idx, column=5).value)  # E
        valor = _safe_float(ws.cell(row=row_idx, column=7).value)  # G

        if not ticker or not transacao:
            continue

        # Filtra apenas COMPRA e VENDA (ignora dividendos, JCP)
        if transacao not in ("COMPRA", "VENDA"):
            continue

        if ticker not in posicoes:
            posicoes[ticker] = {"qtd": 0, "custo_total": 0.0}

        if transacao == "COMPRA":
            posicoes[ticker]["qtd"] += int(qtd)
            posicoes[ticker]["custo_total"] += valor
        elif transacao == "VENDA":
            posicoes[ticker]["qtd"] -= int(qtd)
            # Reduz custo proporcional
            if posicoes[ticker]["qtd"] + int(qtd) > 0:
                custo_medio = posicoes[ticker]["custo_total"] / (posicoes[ticker]["qtd"] + int(qtd))
                posicoes[ticker]["custo_total"] -= custo_medio * int(qtd)
            else:
                posicoes[ticker]["custo_total"] = 0

    ativos = []
    for ticker, pos in posicoes.items():
        if pos["qtd"] <= 0:
            continue

        saldo_bruto = pos["custo_total"]
        if saldo_bruto <= 0:
            saldo_bruto = 0

        tipo = "Acao"
        if ticker.endswith("11") and len(ticker) >= 5:
            tipo = "FII"

        ativos.append(ParsedAssetInst(
            nome=ticker,
            ticker=ticker,
            tipo=tipo,
            subtipo=tipo,
            saldo_bruto=saldo_bruto,
            saldo_liquido=saldo_bruto,
            impostos=0.0,
            alocacao_pct=0.0,
            rent_mes=0.0,
            rent_ano=0.0,
            rent_12m=0.0,
            indexador="",
            taxa="",
            vencimento="",
            emissor="",
            instituicao=instituicao,
        ))

    return ativos


def _parse_sumario(ws) -> dict:
    """Extrai dados do sumario do BTG."""
    result = {"patrimonio_bruto": 0.0, "patrimonio_liquido": 0.0}

    for row_idx in range(6, min(20, ws.max_row + 1)):
        mercado = _safe_str(ws.cell(row=row_idx, column=2).value)  # B
        if mercado.lower() == "total":
            result["patrimonio_bruto"] = _safe_float(ws.cell(row=row_idx, column=5).value)  # E
            result["patrimonio_liquido"] = _safe_float(ws.cell(row=row_idx, column=6).value)  # F
            break

    return result


def _parse_capa(ws) -> dict:
    """Extrai dados da capa do BTG."""
    result = {"cliente": "", "conta": "", "periodo": "", "data_emissao": ""}

    for row_idx in range(1, min(30, ws.max_row + 1)):
        for col_idx in range(1, 6):
            val = _safe_str(ws.cell(row=row_idx, column=col_idx).value)
            if not val:
                continue

            if "Conta Controle" in val:
                m = re.search(r'(\d+)', val)
                if m:
                    result["conta"] = m.group(1)
            elif "CPF" in val:
                pass  # Nao armazenamos CPF
            elif "Período" in val or "Periodo" in val:
                result["periodo"] = val
            elif "Emitido" in val:
                m = re.search(r'(\d{2}/\d{2}/\d{2})', val)
                if m:
                    result["data_emissao"] = m.group(1)
            elif row_idx == 17 and col_idx == 3:
                # Nome do cliente geralmente na row 17, col C
                result["cliente"] = val

    return result


def parse_btg_xlsx(file_path: str) -> InstitutionData:
    """Parseia relatorio BTG Pactual XLSX e retorna InstitutionData."""
    wb = openpyxl.load_workbook(file_path, data_only=True)

    # Capa
    capa = {}
    if "Capa" in wb.sheetnames:
        capa = _parse_capa(wb["Capa"])

    # Sumario
    sumario = {"patrimonio_bruto": 0.0, "patrimonio_liquido": 0.0}
    if "Sumario" in wb.sheetnames:
        sumario = _parse_sumario(wb["Sumario"])

    # Renda Fixa
    ativos_rf = []
    if "Renda Fixa" in wb.sheetnames:
        ativos_rf = _parse_renda_fixa(wb["Renda Fixa"], "BTG Pactual")

    # Renda Variavel
    ativos_rv = []
    if "Renda Variavel" in wb.sheetnames:
        ativos_rv = _parse_renda_variavel(wb["Renda Variavel"], "BTG Pactual")

    all_ativos = ativos_rf + ativos_rv

    # Calcula patrimonio total incluindo RV
    patrimonio_bruto = sumario["patrimonio_bruto"]
    if not patrimonio_bruto:
        patrimonio_bruto = sum(a.saldo_bruto for a in all_ativos)
    patrimonio_liquido = sumario["patrimonio_liquido"]
    if not patrimonio_liquido:
        patrimonio_liquido = sum(a.saldo_liquido for a in all_ativos)

    # Recalcula alocacao
    total = patrimonio_bruto or 1
    for a in all_ativos:
        a.alocacao_pct = a.saldo_bruto / total * 100

    # Distribuicao por tipo
    dist_tipo = {}
    for a in all_ativos:
        if a.tipo not in dist_tipo:
            dist_tipo[a.tipo] = 0.0
        dist_tipo[a.tipo] += a.alocacao_pct

    wb.close()

    return InstitutionData(
        nome="BTG Pactual",
        cliente=capa.get("cliente", ""),
        conta=capa.get("conta", ""),
        data_referencia=capa.get("data_emissao", ""),
        patrimonio_bruto=patrimonio_bruto,
        patrimonio_liquido=patrimonio_liquido,
        impostos_totais=sum(a.impostos for a in all_ativos),
        rent_carteira_mes=0.0,
        rent_carteira_ano=0.0,
        rent_carteira_12m=0.0,
        cdi_mes=0.0,
        cdi_ano=0.0,
        perfil_investidor="",
        ativos=all_ativos,
        distribuicao_tipo=dist_tipo,
    )
